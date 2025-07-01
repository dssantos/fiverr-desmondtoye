import os
import csv
from time import sleep
from datetime import datetime
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from pytubefix import YouTube, Playlist
from pytubefix.cli import on_progress
from pytubefix.exceptions import VideoUnavailable
import boto3
from decouple import config

DOWNLOADS_PATH = 'downloads'

# === Configuração de Logging ===

# Cria diretório de logs (se não existir)
os.makedirs("logs", exist_ok=True)

# Gera nome do arquivo de log com data e hora atual
log_filename = datetime.now().strftime("logs/%Y-%m-%d_%H-%M-%S.log")

# Configuração de logging com arquivo separado por execução
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_filename, mode='a', encoding='utf-8'),
        logging.StreamHandler()
    ],
    force=True 
)

# Nome do arquivo CSV
CSV_FILE = 'links.csv'  # ajuste para o nome do seu arquivo

def update_csv(video_data):
    """Atualiza ou adiciona a linha do CSV com base em video_data['url']."""
    rows = []
    found = False

    # Lê o arquivo se existir
    if os.path.isfile(CSV_FILE):
        with open(CSV_FILE, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                if row.get('url') == video_data['url']:
                    rows.append(video_data)  # substitui linha existente
                    found = True
                else:
                    rows.append(row)

    # Se não achou, adiciona como novo
    if not found:
        rows.append(video_data)

    # Escreve tudo de volta (ou cria novo arquivo com cabeçalho)
    with open(CSV_FILE, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=video_data.keys())
        writer.writeheader()
        writer.writerows(rows)

def extrair_links_com_ids(arquivo_xlsx):
    try:
        # Carrega o arquivo Excel
        wb = load_workbook(arquivo_xlsx)
        planilha = wb.active
        
        # Encontra a coluna 'LINK'
        coluna_link = None
        for cell in planilha[1]:  # Verifica a primeira linha (cabeçalho)
            if cell.value and str(cell.value).strip().upper() == 'LINK':
                coluna_link = cell.column_letter
                break
        
        if not coluna_link:
            raise ValueError("Coluna 'LINK' não encontrada na planilha")
        
        # Extrai os dados
        dados = []
        for idx, row in enumerate(planilha.iter_rows(min_row=2, values_only=True), start=2):
            link = row[ord(coluna_link.lower()) - ord('a')]  # Converte letra para índice
            if link:  # Ignora linhas vazias
                dados.append({'id': idx, 'link': str(link)})
        
        return dados
        
    except Exception as e:
        logging.error(f"Erro ao processar arquivo: {str(e)}")
        return []

def extract_urls_from_playlist(url):
    pl = Playlist(url)

    return pl.video_urls

def get_links(csv_file=CSV_FILE):
    all_urls = set()
    
    if not os.path.exists(csv_file):
        return []
    
    with open(csv_file, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            all_urls.add(row['url'])
    
    return list(all_urls)

def list_metadata(csv_file=CSV_FILE):
    rows = []
    if os.path.isfile(csv_file):
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                video_metadata = {
                    'id': row.get('id'),
                    'url': row.get('url'),
                    'title': row.get('title'),
                    'playlist': row.get('playlist'),
                    'length': row.get('length'),
                    'downloaded': row.get('downloaded')
                }
                rows.append(video_metadata)

    return rows

def retrieve_metadata_from_url(url, csv_file=CSV_FILE):
    if os.path.isfile(csv_file):
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                if row.get('url') == url:
                    video_metadata = {
                        'id': row.get('id'),
                        'url': row.get('url'),
                        'title': row.get('title'),
                        'playlist': row.get('playlist'),
                        'length': row.get('length'),
                        'downloaded': row.get('downloaded')
                    }
                    return video_metadata

def download_video(url):
    # yt = YouTube(url, use_oauth=True, allow_oauth_cache=True, on_progress_callback=on_progress)
    # ys = yt.streams.get_highest_resolution()
    # ys.download(output_path=DOWNLOADS_PATH)

    yt = YouTube(
            url,
            use_oauth=True,
            allow_oauth_cache=True,
            on_progress_callback=lambda stream, chunk, bytes_remaining: 
                logging.info(f"Downloading {url} ... {bytes_remaining / (1024 * 1024):.2f}MB remaining"),
            on_complete_callback=lambda stream, file_path: 
                logging.info(f"Download completed: {file_path}")
        )
    try:
        stream = yt.streams.filter(
            progressive=True,
            file_extension='mp4'
        ).order_by('resolution').desc().first()
    except VideoUnavailable as e:
        logging.warning(f'Video unavailable: {url} {e}')
        return 'FAILED'
    return stream.download(output_path=DOWNLOADS_PATH)

def get_metadata(input_url):
    input_url_id = input_url['id']
    input_url = input_url['link']
    if 'playlist?list' in input_url:
        urls = extract_urls_from_playlist(input_url)
        logging.info(f'Playlist Urls to download: {len(urls)} {input_url}')
        playlist = input_url
    else:
        urls = [input_url]
        playlist = None
    existing_urls = get_links()
    for url in urls:
        if url in existing_urls:
            continue
        yt = YouTube(url, use_oauth=True, allow_oauth_cache=True)
        length = yt.length
        title = yt.title

        result = {
            "id": input_url_id,
            "url": url,
            "title": title,
            "playlist": playlist,
            "length": length,
            "downloaded": False
        }
        logging.info(result)
        logging.info('\n')

        update_csv(result)
        sleep(5)

def list_downloaded_files(folder=DOWNLOADS_PATH):
    try:
        folder = Path(folder).expanduser().absolute()
        
        if not folder.exists():
            raise FileNotFoundError(f"Folder not found: {folder}")
            
        files = []
        for item in folder.iterdir():
            if item.is_file():
                files.append(str(item).split('/')[-1])
        
        return files
        
    except Exception as e:
        print(f"Error to list files: {str(e)}")
        return []

def find_duplicated(csv_file=CSV_FILE):
    try:
        df = pd.read_csv(csv_file)
        if 'title' not in df.columns:
            return []
            
        duplicated = df[df.duplicated('title', keep=False)]['title'].unique()
        return list(duplicated)
        
    except Exception as e:
        logging.info(f"Error: {str(e)}")
        return []

s3 = boto3.client(
    's3',
    aws_access_key_id=config('S3_ACCESS_KEY'),
    aws_secret_access_key=config('S3_SECRET_KEY'),
    region_name=config('S3_REGION')  # Ex: 'us-east-1'
)

def s3_folder_exists(bucket_name, pasta_path):
    if not pasta_path.endswith('/'):
        pasta_path += '/'
    
    response = s3.list_objects_v2(
        Bucket=bucket_name,
        Prefix=pasta_path,
        MaxKeys=1
    )
    
    return 'Contents' in response

def create_s3_folder(bucket_name, folder_path):
    if not folder_path.endswith('/'):
        folder_path += '/'
    
    s3.put_object(
        Bucket=bucket_name,
        Key=folder_path
    )
    logging.info(f"Folder '{folder_path}' created.")

def upload_file_to_s3(bucket_name, local_path, s3_path):
    try:
        s3.upload_file(
            local_path,
            bucket_name,
            s3_path,
            ExtraArgs={
                'ACL': 'bucket-owner-full-control',  # Importante para acesso do cliente
                'Metadata': {
                    'uploaded-by': 'fiverr'
                }
            }
        )
        logging.info(f"File {local_path} send to {s3_path}")
        return True
    except Exception as e:
        logging.info(f"Error upload: {str(e)}")
        return False

option = 6

if option == 1:
    # Get metadata
    urls_to_download = extrair_links_com_ids('Copy of Pregnant Face Dataset.xlsx')
    logging.info(f'Urls to download: {len(urls_to_download)}')
    for new_url in urls_to_download:
        get_metadata(new_url)
    logging.info('\nFinished metadata!')

elif option == 2:
    # Rename files
    downloaded = list_downloaded_files()
    for video_data in list_metadata():
        video_title = video_data['title']
        if f"{video_title}.mp4" in downloaded and video_title not in find_duplicated():
            video_id = video_data['url'].split('=')[-1]
            if video_id not in video_title:
                try:
                    os.rename(f"downloads/{video_title}.mp4", f"downloads/{video_title} ({video_id}).mp4")
                except FileNotFoundError as e:
                    logging.warning(f"Not found {video_title}.mp4")

elif option == 3:
    # Check downloaded
    downloaded = list_downloaded_files()
    for video_data in list_metadata():
        video_title = video_data['title']
        video_id = video_data['url'].split('=')[-1]
        if video_id in video_title:
            if f"{video_title}.mp4" in downloaded:
                video_data['downloaded'] = True
                update_csv(video_data)
        else:
            if f"{video_title} ({video_id}).mp4" in downloaded:
                video_data['downloaded'] = True
                update_csv(video_data)

elif option == 4:
    # Rename Titles
    for video_data in list_metadata():
        video_title = video_data['title']
        video_id = video_data['url'].split('=')[-1]
        video_data['title'] = f"{video_title} ({video_id})"
        update_csv(video_data)

elif option ==5:
    ## Download videos
    for video_data in list_metadata():
        video_title = video_data['title']
        video_id = video_data['url'].split('=')[-1]
        if video_data['downloaded'] == 'False':
            file_path = download_video(video_data['url'])
            if os.path.exists(file_path):
                os.rename(f"downloads/{video_title}.mp4", f"downloads/{video_title} ({video_id}).mp4")
                video_data['downloaded'] = True
                logging.info(f'Saved: {video_data}')
            else:
                logging.info(f'Failed: {video_data}')
            sleep(10)

elif option == 6:
    ## Upload to S3
    jump = False # change to True to continue from last uploaded
    for video_data in list_metadata():
        s3_folder_name = video_data['id']
        youtube_id = video_data['url'].split('=')[-1]
        video_title = video_data['title']
        if jump:
            if f'{s3_folder_name}/{video_title} ({youtube_id})' == 'PASTE_LAST_UPLOADED':
                jump = False 
            continue
        if youtube_id in video_title:
            filename_s3 = f"{video_title}.mp4"
        else:
            filename_s3 = f"{video_title} ({youtube_id}).mp4"
        if not s3_folder_exists(config('S3_BUCKET'), s3_folder_name):
            create_s3_folder(config('S3_BUCKET'), s3_folder_name)
        
        filename_local = f"{video_title} ({youtube_id}).mp4"
        if not upload_file_to_s3(
            config('S3_BUCKET'),
            f'{DOWNLOADS_PATH}/{filename_local}',
            f'{s3_folder_name}/{filename_s3}'
        ):
            filename_local = f"{video_title}.mp4"
            upload_file_to_s3(
                config('S3_BUCKET'),
                f'{DOWNLOADS_PATH}/{filename_local}',
                f'{s3_folder_name}/{filename_s3}')
        sleep(1)
            


